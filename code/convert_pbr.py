"""
Convert FnGuide DataGuide PBR.xlsx (daily PBR, 1989+, KOSPI+KOSDAQ incl. delisted)
to a month-end long parquet:  Code, Date(month-end), PBR.
B/M for HML is later taken as 1/PBR.

DataGuide layout (auto-detected): a header row containing '코드' & '코드명', with
date columns (datetimes) starting a few columns later; each subsequent row is one
(code,item) series.
"""
from pathlib import Path
import time
import numpy as np
import pandas as pd
from openpyxl import load_workbook

SRC = Path("/mnt/20t/study/mom_paper_test/data/PBR.xlsx")
OUT = Path("/mnt/20t/졸업논문/data/processed/pbr_monthly.parquet")
OUT.parent.mkdir(parents=True, exist_ok=True)


def main():
    print(f"[open] {SRC} ({SRC.stat().st_size/1e9:.2f} GB)")
    t0 = time.time()
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    header_i = None; code_c = name_c = item_c = date_c0 = None; dates = None
    rows_iter = ws.iter_rows(values_only=True)
    for i, row in enumerate(rows_iter):
        if header_i is None:
            vals = [str(x) if x is not None else "" for x in row]
            if "코드" in vals and "코드명" in vals:
                code_c = vals.index("코드"); name_c = vals.index("코드명")
                item_c = vals.index("아이템명") if "아이템명" in vals else None
                # first datetime column = start of dates
                for j, x in enumerate(row):
                    if hasattr(x, "year") and not isinstance(x, (int, float)):
                        date_c0 = j; break
                dates = pd.to_datetime([row[j] for j in range(date_c0, len(row))], errors="coerce")
                header_i = i
                print(f"[header] row {i}: code_c={code_c} date_c0={date_c0} "
                      f"ndates={len(dates)} {dates[0].date()}..{dates[-1].date()}")
                # month-end column mask
                dser = pd.Series(range(len(dates)), index=dates)
                me_pos = dser.groupby([dates.year, dates.month]).last().values
                me_dates = dates[me_pos]
                break

    series = {}
    n = 0
    for i, row in enumerate(rows_iter):  # continues after header
        code = row[code_c]
        if code is None:
            continue
        vals = row[date_c0:]
        arr = np.array([vals[p] if p < len(vals) else None for p in me_pos], dtype=object)
        series[str(code)] = arr
        n += 1
        if n % 1000 == 0:
            print(f"  {n} codes, {(time.time()-t0)/60:.1f}min", flush=True)
    wb.close()

    wide = pd.DataFrame.from_dict(series, orient="index", columns=me_dates)
    wide.index.name = "Code"
    long = wide.reset_index().melt(id_vars="Code", var_name="Date", value_name="PBR")
    long["PBR"] = pd.to_numeric(long["PBR"], errors="coerce")
    long["Code"] = long["Code"].astype(str).str.lstrip("A").str.zfill(6)
    long["Date"] = pd.to_datetime(long["Date"])
    long = long.dropna(subset=["PBR"])
    long = long[long["PBR"] > 0]
    long.to_parquet(OUT, index=False)
    print(f"[out] {OUT}  {len(long):,} rows  {long.Code.nunique()} codes  "
          f"{long.Date.min().date()}..{long.Date.max().date()}  ({(time.time()-t0)/60:.1f}min)")


if __name__ == "__main__":
    main()
